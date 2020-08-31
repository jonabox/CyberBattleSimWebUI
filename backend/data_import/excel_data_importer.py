#import data from the excel file

from config import logging,_createInstance
from openpyxl import load_workbook
import json

class ExcelDataImporter(object):
    def load_data_from_excel(self,data_file_name,meta_config_file):
        try:
            logging.info(f'Start to open Excel File storing the data:{data_file_name}')
            wb=load_workbook(filename=data_file_name)
            logging.info(f'Open Excel File storing the data:{data_file_name}')

            meta_data_config = None
            with open(meta_config_file, 'rt') as f:
                meta_data_config = json.load(f)
            logging.info(f'Open Configulation File storing the mapping information:{meta_config_file}')

            data = {}
            if meta_data_config is not None:
                for config_item in meta_data_config['config']:
                    instance_list = self.__load_instance_data_from_excel(wb, config_item['module_name'], config_item['class_name'], config_item['sheet_name'], config_item['row_number'],config_item['column_mapping'],config_item['column_default'])
                    data[config_item['class_name']] = instance_list
            logging.info(f'Finish Data Loading from Excel File:{len(data)}')
            return data
        except Exception as inst:
            logging.error(f'Fail to load data: unknown error:{inst}')
            return None
    
    #wb: excel book for the data
    #module_name: the name of the module
    #class_name: the name of the given instance
    #sheet_name: the name of sheet that have the data
    #row_number: the total number of the data
    #column_mapping: the mapping relation between the column and the attribute of the instance
    #column_default: the mapping relation between the default value and the attribute of the instance
    def __load_instance_data_from_excel(self,wb,module_name, class_name, sheet_name, row_number,column_mapping,column_default):
        data_sheet = wb[sheet_name]
        logging.info(f"start create instance:{class_name}:{row_number}")
        instance_list = []
        for row in range(2,row_number+1):
            #create an instance
            instance = _createInstance(module_name, class_name)

            #for each column, set the value
            for attribute_name,column_index in column_mapping.items():
                attribute_value = data_sheet.cell(row=row,column=column_index).value
                if attribute_value is not None and type(attribute_value) is str:
                    attribute_value = attribute_value.strip()
                setattr(instance,attribute_name,attribute_value)

            #set the default value
            for attribute_name,attribute_default_value in column_default.items():
                setattr(instance,attribute_name,attribute_default_value)
            #need function to remove the duplicated
            instance_list.append(instance)
        logging.info(f'finish load data within sheet:{sheet_name}:{len(instance_list)}')
        return instance_list
